import datetime
from typing import Generator
import aiohttp

import openpyxl.cell
from openpyxl.reader.excel import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from requests import get
from telebot.formatting import hbold

import config
from timetable_utils import ChangedClass


no_changes_text = "Изменений нет."
no_changes_yet_text = "Изменения пока что не сделали."
changes_file_name = "changes.xlsx"


async def get_timetable_changes(date: datetime.datetime) -> str:
    async with aiohttp.ClientSession() as session:
        async with session.get(config.TIMETABLE_CHANGES_URL) as session_response:
            response = session_response

            if response.status == 200:
                with open(changes_file_name, "wb") as changes_table_file:
                    changes_table_file.write(await response.read())
            else:
                print(f"Error while getting changes file. Status code: {response.status}")
                return "Ошибка при получении изменений в расписании на стороне бота. Уже исправляем."

    changes_workbook: Workbook = load_workbook(changes_file_name, read_only=True)
    sheet_name = date_to_sheet_name(date)

    try:
        changes_sheet: Worksheet = changes_workbook[sheet_name]
    except KeyError:
        return hbold(no_changes_yet_text)

    rows_with_changes: list[int] = list()
    rows_iterator: Generator[tuple[openpyxl.cell.Cell], None, None] = changes_sheet.iter_rows(min_row=1, max_row=100,
                                                                                              min_col=2, max_col=2)
    for row in rows_iterator:
        for cell in row:
            if cell.value == config.GROUP_NAME:
                rows_with_changes.append(cell.row)

    if len(rows_with_changes) == 0:
        return hbold(no_changes_text)

    result = hbold("Изменения в расписании:\n\n")

    for row_number in rows_with_changes:
        class_number_was = changes_sheet.cell(row_number, 3).value
        class_room_was = changes_sheet.cell(row_number, 4).value
        class_subject_was = changes_sheet.cell(row_number, 5).value
        class_teacher_was = changes_sheet.cell(row_number, 6).value
        class_number_became = changes_sheet.cell(row_number, 8).value
        class_room_became = changes_sheet.cell(row_number, 9).value
        class_subject_became = changes_sheet.cell(row_number, 10).value
        class_teacher_became = changes_sheet.cell(row_number, 11).value
        changed_class = ChangedClass(class_number_was, class_room_was, class_subject_was, class_teacher_was,
                                     class_number_became, class_room_became, class_subject_became, class_teacher_became)
        result += str(changed_class)
        result += "\n"

    return result


def date_to_sheet_name(date: datetime.datetime):
    return date.strftime("%d.%m")
